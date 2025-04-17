# +----------------------------------------------------------------------
# | DjangoAdmin敏捷开发框架 [ 赋能开发者，助力企业发展 ]
# +----------------------------------------------------------------------
# | 版权所有 2021~2024 北京DjangoAdmin研发中心
# +----------------------------------------------------------------------
# | Licensed LGPL-3.0 DjangoAdmin并不是自由软件，未经许可禁止去掉相关版权
# +----------------------------------------------------------------------
# | 官方网站: https://www.djangoadmin.cn
# +----------------------------------------------------------------------
# | 作者: @一米阳光 团队荣誉出品
# +----------------------------------------------------------------------
# | 版权和免责声明:
# | 本团队对该软件框架产品拥有知识产权（包括但不限于商标权、专利权、著作权、商业秘密等）
# | 均受到相关法律法规的保护，任何个人、组织和单位不得在未经本团队书面授权的情况下对所授权
# | 软件框架产品本身申请相关的知识产权，禁止用于任何违法、侵害他人合法权益等恶意的行为，禁
# | 止用于任何违反我国法律法规的一切项目研发，任何个人、组织和单位用于项目研发而产生的任何
# | 意外、疏忽、合约毁坏、诽谤、版权或知识产权侵犯及其造成的损失 (包括但不限于直接、间接、
# | 附带或衍生的损失等)，本团队不承担任何法律责任，本软件框架禁止任何单位和个人、组织用于
# | 任何违法、侵害他人合法利益等恶意的行为，如有发现违规、违法的犯罪行为，本团队将无条件配
# | 合公安机关调查取证同时保留一切以法律手段起诉的权利，本软件框架只能用于公司和个人内部的
# | 法律所允许的合法合规的软件产品研发，详细声明内容请阅读《框架免责声明》附件；
# +----------------------------------------------------------------------
import os
import random
import time

from flask import request
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from sqlalchemy import and_

import utils
from apps.constants.message import PAGE_LIMIT
from apps.forms.level import LevelForm, LevelStatusForm
from apps.models.level import Level
from apps.services import upload
from config.env import FLASK_TEMP_PATH, FLASK_IMAGE_URL, FLASK_UPLOAD_DIR
from extends import db
from utils import R, regular
from utils.file import mkdir

from utils.utils import uid


# 查询职级分页数据
def LevelList():
    # 页码
    page = int(request.args.get('page', 1))
    # 每页数
    limit = int(request.args.get('limit', PAGE_LIMIT))
    # 查询数据
    query = Level.query.filter(Level.is_delete == 0)
    # 职级名称
    name = request.args.get('name')
    if name:
        query = query.filter(Level.name.like('%' + name + '%'))
    # 职级状态
    status = request.args.get('status')
    if status:
        query = query.filter(status=status)
    # 排序
    query = query.order_by(Level.sort.asc())
    # 记录总数
    count = query.count()
    # 分页查询
    level_list = query.limit(limit).offset((page - 1) * limit).all()
    # 遍历数据源
    result = []
    if len(level_list) > 0:
        for item in level_list:
            # 对象转字典
            data = utils.load2dict(item)
            # 创建时间
            data['create_time'] = str(item.create_time.strftime('%Y-%m-%d %H:%M:%S')) if item.create_time else None
            # 更新时间
            data['update_time'] = str(item.update_time.strftime('%Y-%m-%d %H:%M:%S')) if item.update_time else None
            # 加入列表
            result.append(data)
    # 返回结果
    return R.ok(data=result, count=count)


# 根据ID查询职级详情
def LevelDetail(level_id):
    # 根据ID查询职级
    level = Level.query.filter(and_(Level.id == level_id, Level.is_delete == 0)).first()
    # 查询结果为空判断
    if not level:
        return None
    # 对象转字典
    data = utils.load2dict(level)
    # 返回结果
    return data


# 添加职级
def LevelAdd():
    # 表单验证
    form = LevelForm(request.form)
    if not form.validate():
        # 获取错误描述
        err_msg = regular.get_err(form)
        # 返回错误信息
        return R.failed(msg=err_msg)

    # 表单数据赋值给对象
    level = Level(**form.data)
    level.create_user = uid()
    # 插入数据
    level.save()
    # 返回结果
    return R.ok(msg="添加成功")


# 更新职级
def LevelUpdate():
    # 表单验证
    form = LevelForm(request.form)
    if not form.validate():
        # 获取错误描述
        err_msg = regular.get_err(form)
        # 返回错误信息
        return R.failed(msg=err_msg)

    # 记录ID判空
    id = form.data['id']
    if not id or int(id) <= 0:
        return R.failed("记录ID不能为空")

    # 根据ID查询记录
    level = Level.query.filter(and_(Level.id == id, Level.is_delete == 0)).first()
    # 查询结果判空
    if not level:
        return R.failed("记录不存在")
    result = Level.query.filter_by(id=id).update(form.data)
    # 提交数据
    db.session.commit()
    if not result:
        return R.failed("更新失败")
    # 返回结果
    return R.ok(msg="更新成功")


# 删除职级
def LevelDelete(level_id):
    # 记录ID为空判断
    if not level_id:
        return R.failed("记录ID不存在")
    # 分裂字符串
    list = level_id.split(',')
    # 计数器
    count = 0
    # 遍历数据源
    if len(list) > 0:
        for vId in list:
            # 根据ID查询记录
            level = Level.query.filter(and_(Level.id == int(vId), Level.is_delete == 0)).first()
            # 查询结果判空
            if not level:
                return R.failed("记录不存在")
            # 设置删除标识
            level.is_delete = 1
            # 提交数据
            db.session.commit()
            # 计数器+1
            count += 1
    # 返回结果
    return R.ok(msg="本次共删除{0}条数据".format(count))


# 设置状态
def LevelStatus():
    # 表单验证
    form = LevelStatusForm(request.form)
    if not form.validate():
        # 获取错误描述
        err_msg = regular.get_err(form)
        # 返回错误信息
        return R.failed(msg=err_msg)
    # 记录ID
    id = int(form.data['id'])
    # 状态值
    status = int(form.data['status'])
    # 根据ID查询记录
    level = Level.query.filter(and_(Level.id == id, Level.is_delete == 0)).first()
    # 查询结果判空
    if not level:
        return R.failed("记录不存在")

    # 更新记录
    result = Level.query.filter_by(id=id).update({
        "status": status
    })
    # 提交数据
    db.session.commit()
    if not result:
        return R.failed("设置失败")
    # 返回结果
    return R.ok(msg="设置成功")


# 导入Excel
def LevelImport():
    # 读取上传的文件
    file = upload.uploadFile(".xls|.xlsx")
    # 本地文件地址
    path = file['filePath']
    # 打开工作簿
    wb = load_workbook(path)
    # 获取表单
    ws = wb.worksheets[0]
    # 读取最大行
    max_row = ws.max_row
    # 计数器
    count = 0
    # 遍历读取数据源
    for item in range(2, max_row + 1):
        # 职级名称
        name = ws.cell(item, 2).value
        # 职级排序
        sort = ws.cell(item, 3).value
        # 职级状态
        status = ws.cell(item, 4).value
        # 创建数据
        level = Level(
            id=None,
            name=name,
            status=1 if status == "正常" else 2,
            sort=sort,
        )
        level.create_user = uid()
        # 插入数据
        db.session.add(level)
        db.session.commit()

        count += 1
    # 返回结果
    return R.ok(msg="本次共导入【{}】条数据".format(count))


# 导出Excel
def LevelExport():
    # 查询数据
    query = Level.query.filter(Level.is_delete == 0)
    # 职级名称
    name = request.args.get('name')
    if name:
        query = query.filter(Level.name.like('%' + name + '%'))
    # 职级状态
    status = request.args.get('status')
    if status:
        query = query.filter(status=status)
    # 排序
    query = query.order_by(Level.sort.asc())
    # 查询职级列表
    level_list = query.all()
    # 遍历数据源
    data_list = []
    if len(level_list) > 0:
        for data in level_list:
            data = {
                'id': data.id,
                'name': data.name,
                'sort': data.sort,
                'status': "正常" if data.status == 1 else "禁用",
                'create_time': str(data.create_time.strftime('%Y-%m-%d %H:%M:%S')) if data.create_time else None,
                'update_time': str(data.update_time.strftime('%Y-%m-%d %H:%M:%S')) if data.update_time else None,
            }
            data_list.append(data)

    # 实例化工作薄
    wb = Workbook()
    # 指定第一个sheet
    ws = wb.worksheets[0]
    # 添加表头
    ws.append(['职级ID', '职级名称', '职级排序', '职级状态', '创建时间', '更新时间'])

    # 遍历外层列表
    for data in data_list:
        item_list = []
        # 遍历内层每一个字典dict，把dict每一个值存入list
        for data in data.values():
            item_list.append(data)
        # 加入数据
        ws.append(item_list)

    # 保存文件
    save_path = FLASK_TEMP_PATH + "/" + time.strftime('%Y%m%d')
    # 创建存放目录
    mkdir(save_path)
    # 定义文件名，年月日时分秒随机数
    file_name = time.strftime('%Y%m%d%H%M%S') + '%05d' % random.randint(0, 100)
    # 重写合成文件名
    file_name = os.path.join(file_name + ".xlsx")
    # 拼接文件路径
    path = save_path + "/" + file_name
    # 写入文件
    wb.save(r'' + path + '')
    # 文件访问地址
    file_url = FLASK_IMAGE_URL + path.replace(FLASK_UPLOAD_DIR, "")
    # 返回结果
    return R.ok(data=file_url)


# 获取职级数据列表
def getLevelList():
    # 查询职级数据列表
    list = Level.query.filter(and_(Level.is_delete == 0, Level.status == 1)).order_by(Level.sort.asc()).all()
    # 实例化列表
    level_list = []
    # 遍历数据源
    if list:
        for v in list:
            # 对象转字典
            item = utils.load2dict(v)
            # 加入数组
            level_list.append(item)
    # 返回结果
    return level_list
